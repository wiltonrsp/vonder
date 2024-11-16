from openpyxl import Workbook, load_workbook
import configparser
import mysql.connector
import urllib.request
import os
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager


#configuracoes
config = configparser.ConfigParser()
config.read("config2.ini")

#conexao banco de dados
conexao = mysql.connector.connect(
            host=config.get("Credenciais", "bd_ip"),
            port=3306,
            user=config.get("Credenciais", "bd_user"),
            database=config.get("Credenciais", "bd_db"),
            password=config.get("Credenciais", "bd_password"))
cursor = conexao.cursor()

comando = (
           "CREATE TABLE IF NOT EXISTS produtos ( id_produto VARCHAR(15));"
)
cursor.execute(comando)

conexao.commit()
#xlsx
tabela = load_workbook(config.get("Credenciais", "xlsx"))
tabela = tabela.worksheets[0]

config.get("Credenciais", "xlsx")
#criando arquivo de saida
wb = Workbook()
ws = wb.active
ws.title = "Retorno"
 

#navegador
option = webdriver.ChromeOptions()
option.add_argument("start-maximized")
option.add_argument("--disable-infobars")
option.add_argument("--disable-extensions")
option.add_argument("--disable-popup-block")
option.add_argument("no-default-browser-check")
option.add_argument(r"--user-data-dir="+str(config.get("Credenciais", "user_pref"))) 
option.add_argument("--profile-directory=Default")
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()),options=option)
driver.get(config.get("Credenciais", "site"))
kk=0
for line in tabela:
    id_product = (line[0].value)
    comando = "SELECT COUNT(*) FROM produtos WHERE id_produto = '"+str(id_product)+"'"
    cursor.execute(comando)
    
    quantidade_registros = cursor.fetchone()[0]
    conexao.commit()
    if(quantidade_registros==0):

            
        print('Iniciando Extração do produto '+str(id_product))
        
        print('Validando Popup')
        try:
            popup  = driver.find_element(By.XPATH,"//button[@class='campanha-popup-close']")
            popup.click()
        except:
            print('S/n Popup')
        time.sleep(2)
        busca  = driver.find_element(By.XPATH,"//input[@id='buscaTop']")
        busca.click()
        busca.send_keys(id_product)
        print('Buscando....')
        
        
        botaoBusca = driver.find_element(By.XPATH,"//a[@class='botaoBuscar']")
        botaoBusca.click()
        
        time.sleep(5)
        produto = driver.find_element(By.XPATH,"//div[@class='nomeProd']//a")
        produto.click()
        
        print('Coletando')
        categorias = driver.find_elements(By.XPATH,"//div[@class='breadCrumb']//span")
        j=0
        categorias_final = ''
        for i in categorias:
            if(j!=0):
               
               auxiliar=i.text
               auxiliar=auxiliar.replace('| ','')
               if(j<len(categorias)-1):
                   categorias_final = categorias_final+auxiliar+'|'
               else:
                   categorias_final = categorias_final+auxiliar
               j+=1
            else:
               j+=1

        nome_produto   = driver.find_element(By.XPATH,"//h1[@class='nomeProduto']")
        nome_produto   = nome_produto.get_attribute('innerHTML')
        nome_produto   = nome_produto.split('<label')
        nome_produto   = nome_produto[0]
        
        nome_produto   = nome_produto.replace('\n','')
        nome_produto   = nome_produto.replace('  ','')
        nome_produto_final   = nome_produto.strip()
        
        codigo_produto = driver.find_element(By.XPATH,"//label[@id='codigoProduto']")
        codigo_produto = codigo_produto.get_attribute('innerHTML')
        codigo_produto = codigo_produto.replace('\n','')
        codigo_produto_final = codigo_produto.strip()
        descricaoProd  = driver.find_element(By.XPATH,"//div[@class='descricaoProd']")
        auxiliar_html = descricaoProd.get_attribute('innerHTML')
 
        auxiliar = descricaoProd.text
        auxiliar = auxiliar.split('Certificados:')
        auxiliar_html = auxiliar_html.split('<b>Certificados: </b>')
        certificados = ''
        conteudo = ''
        conteudo_html = ''
        certificados_html = ''
        
        #captura dos certificados e conteudos
        if(len(auxiliar)>1):
            conteudo_html    = auxiliar_html[0]
            conteudo     = auxiliar[0]
            certificados = auxiliar[1]
            certificados_html = auxiliar_html[1]
            certificados = certificados.split('Garantia legal:')
            certificados_html = certificados_html.split('<p style="margin-left: 5px; padding-bottom: 10px;">Garantia legal:')
            certificados = certificados[0]
            certificados_html = certificados_html[0]
            conteudo     = conteudo.replace('\n','')
            certificados = certificados.replace('\n','')
        else:
            conteudo_html = auxiliar_html[0]
            conteudo = auxiliar[0]
            conteudo = conteudo.split('Garantia legal:')
            conteudo_html = conteudo_html.split('<p style="margin-left: 5px; padding-bottom: 10px;">Garantia legal:')
            conteudo = conteudo[0]
            conteudo_html = conteudo_html[0]
            conteudo =  conteudo.replace('\n','')
 
        #captura das imagens
        try:
            imagens = driver.find_elements(By.XPATH,"//li[@class='selected over']//a//img")
            download_images = config.get("Credenciais", "salvar_imagens")
            imagens_links = ''
            path = config.get('Credenciais','imagens_path')
            for i in imagens:
                link = i.get_attribute("src")
                link = link.replace('https','http')
                nome_link = link
                nome_link = nome_link.split('.jpg')
                nome_link = nome_link[0]
                nome_link = nome_link.split('temp/')
                nome_link = nome_link[1]
                
                if(download_images=='S'):
                    if not os.path.exists(path+'\\'+str(id_product)):
                        os.makedirs(path+'\\'+str(id_product))
                    urllib.request.urlretrieve(link, path+'\\'+str(id_product)+'\\'+nome_link+'.jpg')
                imagens_links=imagens_links+link+'|'
                
        except:
            imagem = driver.find_element(By.XPATH,"//li[@class='fotoGrandeProd']//a//img")
            download_images = config.get("Credenciais", "salvar_imagens")
            imagens_links = ''
            path = config.get('Credenciais','imagens_path')
            link = imagem.get_attribute("src")
            link = link.replace('https','http')
            nome_img = link.split('.jpg')
            nome_img = nome_img[0]
            nome_img = nome_img.split('temp/')
            nome_img = nome_img[1]
            imagens_links = link
            if(download_images=='S'):
                
                if not os.path.exists(path+'\\'+str(id_product)):
                    os.makedirs(path+'\\'+str(id_product))
                    
                urllib.request.urlretrieve(link, path+'\\'+str(id_product)+'\\'+nome_img+'.jpg')
        detalhes = driver.find_element(By.XPATH,"//div[@id='navInfoProdutos']//div[@class='boxAba']//div[@class='content']")     
        url_atual = driver.current_url
        ws.append([str(id_product),str(url_atual),str(codigo_produto_final),str(nome_produto_final),str(imagens_links),str(conteudo),str(conteudo_html),str(detalhes.text),str(detalhes.get_attribute('innerHTML')),str(certificados),str(certificados_html)])
        wb.save(config.get("Credenciais", "xlsx_file"))
        comando = "INSERT INTO produtos (id_produto) VALUES ('"+str(id_product)+"');"
        cursor.execute(comando)
        conexao.commit()
        if(kk==4):
            break;
        else:
            kk+=1
    else:
        print('Produto '+str(id_product)+'já foi extraido')
    










